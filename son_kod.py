# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'arayuz.ui'
#
# Created by: PyQt5 UI code generator 5.14.1
#
# WARNING! All changes made in this file will be lost!


import pickle

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QFileDialog
from openpyxl import load_workbook


def ogrenci_olustur(satir_no, numara, ad, biyoloji, cografya, fizik, kimya, matematik, tarih, edebiyat):
    ogrenci = dict()
    ogrenci["satir"] = satir_no
    ogrenci["numara"] = numara
    ogrenci["ad"] = ad
    if biyoloji is None:
        ogrenci["biyoloji"] = ""
        biyoloji = ""
    elif biyoloji == "" or biyoloji[-18:] == "(Kursa Yerleşmedi)":
        ogrenci["biyoloji"] = ""
    else:
        if biyoloji[-2] == " ":
            if biyoloji[-4].isdigit():
                ogrenci["biyoloji"] = biyoloji[-4] + biyoloji[-3] + biyoloji[-1]
            else:
                ogrenci["biyoloji"] = biyoloji[-3] + biyoloji[-1]
        else:
            ogrenci["biyoloji"] = biyoloji[-5] + biyoloji[-4] + biyoloji[-2] + biyoloji[-1]

    if cografya is None:
        ogrenci["cografya"] = ""
        cografya = ""
    elif cografya == "" or cografya[-18:] == "(Kursa Yerleşmedi)":
        ogrenci["cografya"] = ""
    else:
        if cografya[-2] == " ":
            if cografya[-4].isdigit():
                ogrenci["cografya"] = cografya[-4] + cografya[-3] + cografya[-1]
            else:
                ogrenci["biyoloji"] = biyoloji[-3] + biyoloji[-1]
        else:
            ogrenci["cografya"] = cografya[-5] + cografya[-4] + cografya[-2] + cografya[-1]

    if fizik is None:
        ogrenci["fizik"] = ""
        fizik = ""
    elif fizik == "" or fizik[-18:] == "(Kursa Yerleşmedi)":
        ogrenci["fizik"] = ""
    else:
        if fizik[-2] == " ":
            if fizik[-4].isdigit():
                ogrenci["fizik"] = fizik[-4] + fizik[-3] + fizik[-1]
            else:
                ogrenci["fizik"] = fizik[-3] + fizik[-1]
        else:
            ogrenci["fizik"] = fizik[-5] + fizik[-4] + fizik[-2] + fizik[-1]

    if kimya is None:
        ogrenci["kimya"] = ""
        kimya = ""
    elif kimya == "" or kimya[-18:] == "(Kursa Yerleşmedi)":
        ogrenci["kimya"] = ""
    else:
        if kimya[-2] == " ":
            if kimya[-4].isdigit():
                ogrenci["kimya"] = kimya[-4] + kimya[-3] + kimya[-1]
            else:
                ogrenci["kimya"] = kimya[-3] + kimya[-1]
        else:
            ogrenci["kimya"] = kimya[-5] + kimya[-4] + kimya[-2] + kimya[-1]

    if matematik is None:
        ogrenci["matematik"] = ""
        matematik = ""
    elif matematik == "" or matematik[-18:] == "(Kursa Yerleşmedi)":
        ogrenci["matematik"] = ""
    else:
        if matematik[-2] == " ":
            if matematik[-4].isdigit():
                ogrenci["matematik"] = matematik[-4] + matematik[-3] + matematik[-1]
            else:
                ogrenci["matematik"] = matematik[-3] + matematik[-1]
        else:
            ogrenci["matematik"] = matematik[-5] + matematik[-4] + matematik[-2] + matematik[-1]

    if tarih is None:
        ogrenci["tarih"] = ""
        tarih = ""
    elif tarih == "" or tarih[-18:] == "(Kursa Yerleşmedi)":
        ogrenci["tarih"] = ""
    else:
        if tarih[-2] == " ":
            if tarih[-4].isdigit():
                ogrenci["tarih"] = tarih[-4] + tarih[-3] + tarih[-1]
            else:
                ogrenci["tarih"] = tarih[-3] + tarih[-1]
        else:
            ogrenci["tarih"] = tarih[-5] + tarih[-4] + tarih[-2] + tarih[-1]

    if edebiyat is None:
        ogrenci["edebiyat"] = ""
        edebiyat = ""
    elif edebiyat == "" or edebiyat[-18:] == "(Kursa Yerleşmedi)":
        ogrenci["edebiyat"] = ""
    else:
        if edebiyat[-2] == " ":
            if edebiyat[-4].isdigit():
                ogrenci["edebiyat"] = edebiyat[-4] + edebiyat[-3] + edebiyat[-1]
            else:
                ogrenci["edebiyat"] = edebiyat[-3] + edebiyat[-1]
        else:
            ogrenci["edebiyat"] = edebiyat[-5] + edebiyat[-4] + edebiyat[-2] + edebiyat[-1]

    ogrenci["biyoloji_ogr"] = biyoloji.split("\n")[0][8:]
    ogrenci["cografya_ogr"] = cografya.split("\n")[0][8:]
    ogrenci["fizik_ogr"] = fizik.split("\n")[0][8:]
    ogrenci["kimya_ogr"] = kimya.split("\n")[0][8:]
    ogrenci["matematik_ogr"] = matematik.split("\n")[0][8:]
    ogrenci["tarih_ogr"] = tarih.split("\n")[0][8:]
    ogrenci["edebiyat_ogr"] = edebiyat.split("\n")[0][8:]

    return ogrenci


numara_col = "A"
ad_col = "C"
biyoloji_col = "F"
cografya_col = "G"
fizik_col = "H"
kimya_col = "I"
matematik_col = "J"
tarih_col = "K"
edebiyat_col = "L"


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(800, 525)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.load_lists_button = QtWidgets.QPushButton(self.centralwidget)
        self.load_lists_button.setGeometry(QtCore.QRect(30, 30, 140, 32))
        self.load_lists_button.setObjectName("load_lists_button")
        self.selected_file_text = QtWidgets.QLabel(self.centralwidget)
        self.selected_file_text.setGeometry(QtCore.QRect(175, 30, 320, 32))
        font = QtGui.QFont()
        font.setPointSize(8)
        self.selected_file_text.setFont(font)
        self.selected_file_text.setText("")
        self.selected_file_text.setObjectName("selected_file_text")
        self.create_list_button = QtWidgets.QPushButton(self.centralwidget)
        self.create_list_button.setGeometry(QtCore.QRect(30, 140, 150, 28))
        self.create_list_button.setObjectName("create_list_button")
        self.list_widget = QtWidgets.QListWidget(self.centralwidget)
        self.list_widget.setGeometry(QtCore.QRect(500, 70, 256, 421))
        self.list_widget.setObjectName("list_widget")
        self.okul_adi_text_2 = QtWidgets.QLineEdit(self.centralwidget)
        self.okul_adi_text_2.setGeometry(QtCore.QRect(100, 90, 251, 22))
        self.okul_adi_text_2.setObjectName("okul_adi_text_2")
        self.okul_adi_text = QtWidgets.QLabel(self.centralwidget)
        self.okul_adi_text.setGeometry(QtCore.QRect(30, 90, 55, 22))
        self.okul_adi_text.setObjectName("okul_adi_text")
        self.save_list_button = QtWidgets.QPushButton(self.centralwidget)
        self.save_list_button.setGeometry(QtCore.QRect(30, 200, 150, 28))
        self.save_list_button.setObjectName("save_list_button")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(500, 20, 251, 22))
        self.label.setObjectName("label")
        MainWindow.setCentralWidget(self.centralwidget)

        self.load_lists_button.clicked.connect(self.yukle)
        self.create_list_button.clicked.connect(self.olustur)
        self.save_list_button.clicked.connect(self.kaydet)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
        self.file_name = ""

    def yukle(self):
        self.file_name, _ = QFileDialog.getOpenFileName(filter="Excell Dosyası (*.xlsx)")
        if self.file_name:
            self.selected_file_text.setText(self.file_name)

    def olustur(self):
        if self.file_name:
            wb = load_workbook(filename=self.file_name)
            ws = wb[wb.sheetnames[0]]

            ogrenciler = list()
            numara = None
            s = 2
            while numara is None or numara == "":
                numara_cell = "{}{}".format(numara_col, s)
                numara = ws[numara_cell].value
                s += 1

            for row in range(s-1, ws.max_row + 1):
                # ws[cell_name].value

                numara_cell = "{}{}".format(numara_col, row)
                numara = ws[numara_cell].value

                ad_cell = "{}{}".format(ad_col, row)
                ad = ws[ad_cell].value

                biyoloji_cell = "{}{}".format(biyoloji_col, row)
                biyoloji = ws[biyoloji_cell].value

                cografya_cell = "{}{}".format(cografya_col, row)
                cografya = ws[cografya_cell].value

                fizik_cell = "{}{}".format(fizik_col, row)
                fizik = ws[fizik_cell].value

                kimya_cell = "{}{}".format(kimya_col, row)
                kimya = ws[kimya_cell].value

                matematik_cell = "{}{}".format(matematik_col, row)
                matematik = ws[matematik_cell].value

                tarih_cell = "{}{}".format(tarih_col, row)
                tarih = ws[tarih_cell].value

                edebiyat_cell = "{}{}".format(edebiyat_col, row)
                edebiyat = ws[edebiyat_cell].value
                if edebiyat is None:
                    edebiyat = ""

                ogrenciler.append(
                    ogrenci_olustur(row, numara, ad, biyoloji, cografya, fizik, kimya, matematik, tarih,
                                    edebiyat=edebiyat))

            ogrenciler = sorted(ogrenciler, key=lambda i: i['numara'])
            olustur = list()
            for ogrenci in ogrenciler:
                if ogrenci["biyoloji"]:
                    if "biyoloji" + ogrenci["biyoloji"] not in olustur:
                        olustur.append("biyoloji" + ogrenci["biyoloji"])
                if ogrenci["cografya"]:
                    if "cografya" + ogrenci["cografya"] not in olustur:
                        olustur.append("cografya" + ogrenci["cografya"])
                if ogrenci["fizik"]:
                    if "fizik" + ogrenci["fizik"] not in olustur:
                        olustur.append("fizik" + ogrenci["fizik"])
                if ogrenci["fizik"]:
                    if "fizik" + ogrenci["fizik"] not in olustur:
                        olustur.append("fizik" + ogrenci["fizik"])
                if ogrenci["kimya"]:
                    if "kimya" + ogrenci["kimya"] not in olustur:
                        olustur.append("kimya" + ogrenci["kimya"])
                if ogrenci["matematik"]:
                    if "matematik" + ogrenci["matematik"] not in olustur:
                        olustur.append("matematik" + ogrenci["matematik"])
                if ogrenci["tarih"]:
                    if "tarih" + ogrenci["tarih"] not in olustur:
                        olustur.append("tarih" + ogrenci["tarih"])
                if ogrenci["edebiyat"]:
                    if "edebiyat" + ogrenci["edebiyat"] not in olustur:
                        olustur.append("edebiyat" + ogrenci["edebiyat"])
            self.ekle(olustur)
        else:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Warning)
            msg.setText("Lütfen tercih listesini seçiniz!")
            msg.setWindowTitle(" ")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()

    def kaydet(self):
        okul = self.okul_adi_text_2.text()
        if okul:
            if self.file_name:
                wb = load_workbook(filename=self.file_name)
                ws = wb[wb.sheetnames[0]]

                ogrenciler = list()
                numara = None
                s = 2
                while numara is None or numara == "":
                    numara_cell = "{}{}".format(numara_col, s)
                    numara = ws[numara_cell].value
                    s += 1

                for row in range(s-1, ws.max_row + 1):
                    # ws[cell_name].value
                    numara_cell = "{}{}".format(numara_col, row)
                    numara = ws[numara_cell].value


                    ad_cell = "{}{}".format(ad_col, row)
                    ad = ws[ad_cell].value

                    biyoloji_cell = "{}{}".format(biyoloji_col, row)
                    biyoloji = ws[biyoloji_cell].value

                    cografya_cell = "{}{}".format(cografya_col, row)
                    cografya = ws[cografya_cell].value

                    fizik_cell = "{}{}".format(fizik_col, row)
                    fizik = ws[fizik_cell].value

                    kimya_cell = "{}{}".format(kimya_col, row)
                    kimya = ws[kimya_cell].value

                    matematik_cell = "{}{}".format(matematik_col, row)
                    matematik = ws[matematik_cell].value

                    tarih_cell = "{}{}".format(tarih_col, row)
                    tarih = ws[tarih_cell].value

                    edebiyat_cell = "{}{}".format(edebiyat_col, row)
                    edebiyat = ws[edebiyat_cell].value
                    if edebiyat is None:
                        edebiyat = ""

                    ogrenciler.append(
                        ogrenci_olustur(row, numara, ad, biyoloji, cografya, fizik, kimya, matematik, tarih,
                                        edebiyat=edebiyat))

                ogrenciler = sorted(ogrenciler, key=lambda a: a['numara'])
                olustur = list()
                for ogrenci in ogrenciler:
                    if ogrenci["biyoloji"]:
                        if "biyoloji" + ogrenci["biyoloji"] not in olustur:
                            olustur.append("biyoloji" + ogrenci["biyoloji"])
                    if ogrenci["cografya"]:
                        if "cografya" + ogrenci["cografya"] not in olustur:
                            olustur.append("cografya" + ogrenci["cografya"])
                    if ogrenci["fizik"]:
                        if "fizik" + ogrenci["fizik"] not in olustur:
                            olustur.append("fizik" + ogrenci["fizik"])
                    if ogrenci["fizik"]:
                        if "fizik" + ogrenci["fizik"] not in olustur:
                            olustur.append("fizik" + ogrenci["fizik"])
                    if ogrenci["kimya"]:
                        if "kimya" + ogrenci["kimya"] not in olustur:
                            olustur.append("kimya" + ogrenci["kimya"])
                    if ogrenci["matematik"]:
                        if "matematik" + ogrenci["matematik"] not in olustur:
                            olustur.append("matematik" + ogrenci["matematik"])
                    if ogrenci["tarih"]:
                        if "tarih" + ogrenci["tarih"] not in olustur:
                            olustur.append("tarih" + ogrenci["tarih"])
                    if ogrenci["edebiyat"]:
                        if "edebiyat" + ogrenci["edebiyat"] not in olustur:
                            olustur.append("edebiyat" + ogrenci["edebiyat"])
                self.ekle(olustur)

                kurs_wb = pickle.load(open('lib.dll', 'rb'))

                for i in range(len(olustur) - 1):
                    kurs_wb.copy_worksheet(kurs_wb.active)

                i = 0
                for sheet in kurs_wb:
                    sheet.title = olustur[i]
                    i += 1

                import time

                ay = int(time.strftime("%m"))
                yil = time.strftime("%Y")
                if ay > 8:
                    donem = yil + str(int(yil) + 1)
                else:
                    donem = str(int(yil) - 1) + " - " + yil

                if 2 <= ay <= 6:
                    donem_2 = "2. DÖNEM"
                elif ay == 1 or 9 >= ay:
                    donem_2 = "1. DÖNEM"
                elif 7 == ay or 8 == ay:
                    donem_2 = "YAZ DÖNEMİ"
                else:
                    donem_2 = " "

                okul = self.okul_adi_text_2.text()
                baslik = okul + " " + donem + " EĞİTİM ÖĞRETİM YILI {} SINIFI {} DERSİ " + donem_2 + " KURS YOKLAMA LİSTESİ"

                for sheet in kurs_wb:
                    for k in range(len(sheet.title)):
                        if sheet.title[k].isdigit():
                            ders = sheet.title[:k]
                            break

                    for ogrenci in ogrenciler:
                        k = 5
                        if ogrenci[ders]:
                            if ders + ogrenci[ders] == sheet.title:
                                sheet["D2"].value = ogrenci[ders + "_ogr"]
                                sheet["A1"].value = baslik.format(ogrenci[ders], ders.upper())
                                if sheet["B" + str(k)].value is None:
                                    sheet["B" + str(k)].value = ogrenci["numara"]
                                    sheet["C" + str(k)].value = ogrenci["ad"]
                                else:
                                    while sheet["B" + str(k)].value is not None:
                                        k += 1
                                    sheet["B" + str(k)].value = ogrenci["numara"]
                                    sheet["C" + str(k)].value = ogrenci["ad"]

                self.save_name, _ = QFileDialog.getSaveFileName(filter="Excell Dosyası (*.xlsx)")
                if self.save_name:
                    kurs_wb.save(self.save_name)

            else:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Warning)
                msg.setText("Lütfen tercih listesini seçiniz!")
                msg.setWindowTitle(" ")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
        else:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Warning)
            msg.setText("Lütfen okul adını giriniz!")
            msg.setWindowTitle(" ")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()

    # listeye ekleme
    def ekle(self, items):
        self.list_widget.clear()
        new_items = list()
        for item in items:
            ders, sinif, sube = "", "", ""
            if item[-2].isalpha():
                if item[-4].isdigit():
                    ders, sinif, sube = item[:-4].upper(), item[-4:-2], item[-2:]
            elif item[-2].isdigit():
                if item[-3].isdigit():
                    ders, sinif, sube = item[:-3].upper(), item[-3:-1], item[-1:]
                elif item[-3].isalpha():
                    ders, sinif, sube = item[:-2].upper(), item[-2], item[-1:]

            new_item = "{}/{} {}".format(sinif, sube, ders)
            new_items.append(new_item)
        new_items.sort()
        self.list_widget.addItems(new_items)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.load_lists_button.setText(_translate("MainWindow", "Öğrenci Listesi Yükle"))
        self.create_list_button.setText(_translate("MainWindow", "Yoklamaları Oluştur"))
        self.okul_adi_text.setText(_translate("MainWindow", "Okul Adı:"))
        self.save_list_button.setText(_translate("MainWindow", "Yoklamaları Kaydet"))
        self.label.setText(_translate("MainWindow", "Oluşturulacak Yoklama Listeleri"))


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
