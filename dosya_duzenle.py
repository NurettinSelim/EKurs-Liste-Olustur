import pickle

from openpyxl import load_workbook

import term
from models import Student

cols_dict = dict(numara="A",
                 ad="C",
                 biyoloji="F",
                 cografya="G",
                 fizik="H",
                 kimya="I",
                 matematik="J",
                 tarih="K",
                 edebiyat="L")
excel_ogr_list = "excels/yeni_liste.xlsx"
wb = load_workbook(filename=excel_ogr_list)
ws = wb[wb.sheetnames[0]]
ogrenciler = list()

for row in range(3, ws.max_row + 1):
    # ws[cell_name].value
    data = dict()
    data["satir"] = row
    for col_name, col_id in cols_dict.items():
        data[col_name] = ws[f"{col_id}{row}"].value

    ogrenciler.append(Student(data))

ogrenciler = sorted(ogrenciler, key=lambda i: i.numara)

olustur_set = set()

for ogrenci in ogrenciler:
    for lesson_name, lesson_class in ogrenci.lessons.items():
        if lesson_class:
            olustur_set.add(lesson_name + lesson_class)

olustur_list = list(olustur_set)

# DOSYA OLUŞTUR
kurs_wb = pickle.load(open('lib.dll', 'rb'))
# kurs_wb = load_workbook('sablon.xlsx')

for i in range(len(olustur_list) - 1):
    kurs_wb.copy_worksheet(kurs_wb.active)

for number, sheet in enumerate(kurs_wb):
    sheet.title = olustur_list[number]

baslik = "{} {} EĞİTİM ÖĞRETİM YILI {} SINIFI {} DERSİ {} KURS YOKLAMA LİSTESİ"
okul = " "

for sheet in kurs_wb:
    for k in range(len(sheet.title)):
        if sheet.title[k].isdigit():
            ders = sheet.title[:k]
            break
    for ogrenci in ogrenciler:
        k = 5
        if ders + ogrenci.lesson_from_name(ders).name == sheet.title:
            sheet["D2"].value = ogrenci.lesson_from_name(ders).teacher
            sheet["A1"].value = baslik.format(okul,
                                              term.term_year(),
                                              ogrenci.lesson_from_name(ders).name,
                                              ders.upper(),
                                              term.term_year())
            if sheet["B" + str(k)].value is None:
                sheet["B" + str(k)].value = ogrenci.numara
                sheet["C" + str(k)].value = ogrenci.ad
            else:
                while sheet["B" + str(k)].value:
                    k += 1
                sheet["B" + str(k)].value = ogrenci.numara
                sheet["C" + str(k)].value = ogrenci.ad

kurs_wb.save('excels/listeler.xlsx')
