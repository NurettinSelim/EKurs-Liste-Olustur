from openpyxl import load_workbook


def ogrenci_olustur(satir_no, numara, ad, biyoloji, cografya, fizik, kimya, matematik, edebiyat):
    ogrenci = dict()
    ogrenci["satir"] = satir_no
    ogrenci["numara"] = numara
    ogrenci["ad"] = ad

    if biyoloji == "" or biyoloji[-18:] == "(Kursa Yerleşmedi)":
        ogrenci["biyoloji"] = ""
    else:
        ogrenci["biyoloji"] = biyoloji[-3] + biyoloji[-1]

    if cografya == "" or cografya[-18:] == "(Kursa Yerleşmedi)":
        ogrenci["cografya"] = ""
    else:
        ogrenci["cografya"] = cografya[-3] + cografya[-1]

    if fizik == "" or fizik[-18:] == "(Kursa Yerleşmedi)":
        ogrenci["fizik"] = ""
    else:
        ogrenci["fizik"] = fizik[-3] + fizik[-1]

    if kimya == "" or kimya[-18:] == "(Kursa Yerleşmedi)":
        ogrenci["kimya"] = ""
    else:
        ogrenci["kimya"] = kimya[-3] + kimya[-1]

    if matematik == "" or matematik[-18:] == "(Kursa Yerleşmedi)":
        ogrenci["matematik"] = ""
    else:
        ogrenci["matematik"] = matematik[-3] + matematik[-1]

    if edebiyat == "" or edebiyat[-18:] == "(Kursa Yerleşmedi)":
        ogrenci["edebiyat"] = ""
    else:
        ogrenci["edebiyat"] = edebiyat[-3] + edebiyat[-1]

    return ogrenci


def main(dosya):
    wb = load_workbook(filename=dosya)
    ws = wb['9.Sınıf']

    numara_col = "A"
    ad_col = "C"
    okul_col = "D"
    biyoloji_col = "F"
    cografya_col = "G"
    fizik_col = "H"
    kimya_col = "I"
    matematik_col = "J"
    edebiyat_col = "K"

    ogrenciler = list()

    for row in range(3, ws.max_row + 1):
        # ws[cell_name].value
        okul_cell = "{}{}".format(okul_col, row)
        okul = ws[okul_cell].value
        if okul != "Diğer Okul":
            numara_cell = "{}{}".format(numara_col, row)
            numara = ws[numara_cell].value

            ad_cell = "{}{}".format(ad_col, row)
            ad = ws[ad_cell].value

            biyoloji_cell = "{}{}".format(biyoloji_col, row)
            biyoloji = ws[biyoloji_cell].value

            cografya_cell = "{}{}".format(cografya_col, row)
            cografya = ws[cografya_cell].value

            fizik_cell = "{}{}".format(fizik_col, row)
            fizik = ws[fizik_cell].value

            kimya_cell = "{}{}".format(kimya_col, row)
            kimya = ws[kimya_cell].value

            matematik_cell = "{}{}".format(matematik_col, row)
            matematik = ws[matematik_cell].value

            edebiyat_cell = "{}{}".format(edebiyat_col, row)
            edebiyat = ws[edebiyat_cell].value

            ogrenciler.append(ogrenci_olustur(row, numara, ad, biyoloji, cografya, fizik, kimya, matematik, edebiyat))

    ogrenciler = sorted(ogrenciler, key=lambda i: i['numara'])
    return ogrenciler


if __name__ == '__main__':
    ogrenciler = main('ana_liste.xlsx')
    print(ogrenciler)
