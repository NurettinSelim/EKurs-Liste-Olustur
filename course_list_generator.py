import pickle
import string

from openpyxl import load_workbook

import term
from models import Student


def create_list(filename):
    cols_dict = dict(numara="A", ad="C")
    cols_lesson_dict = dict()
    wb = load_workbook(filename=filename)
    ws = wb[wb.sheetnames[0]]

    for harf in string.ascii_uppercase[5:]:
        if bool(ws[f"{harf}1"].value) == 0:
            break
        cols_lesson_dict[ws[f"{harf}1"].value.lower()] = harf

    students = list()

    for row in range(3, ws.max_row + 1):
        data = dict()
        data["satir"] = row

        for col_name, col_id in cols_lesson_dict.items():
            data[col_name] = ws[f"{col_id}{row}"].value

        for col_name, col_id in cols_dict.items():
            data[col_name] = ws[f"{col_id}{row}"].value

        students.append(Student(data, cols_lesson_dict))

    students = sorted(students, key=lambda i: i.numara)

    olustur_set = set()
    for student in students:
        for lesson_name, lesson_class in student:
            if lesson_class:
                olustur_set.add(lesson_name + lesson_class)

    return dict(lists=sorted(list(olustur_set)),
                students=students)


def save_list_to_excel(filename, schoolname, data_dict):
    lists = data_dict["lists"]
    students = data_dict["students"]

    kurs_wb = pickle.load(open('lib.dll', 'rb'))

    for i in range(len(lists) - 1):
        kurs_wb.copy_worksheet(kurs_wb.active)

    for number, sheet in enumerate(kurs_wb):
        sheet.title = lists[number]

    baslik = "{} {} EĞİTİM ÖĞRETİM YILI {} SINIFI {} DERSİ {} KURS YOKLAMA LİSTESİ"

    for sheet in kurs_wb:
        for k in range(len(sheet.title)):
            if sheet.title[k].isdigit():
                lesson_name = sheet.title[:k]
                break
        for student in students:
            k = 5
            if lesson_name + student.lesson_from_name(lesson_name).name == sheet.title:
                sheet["D2"].value = student.lesson_from_name(lesson_name).teacher
                sheet["A1"].value = baslik.format(schoolname,
                                                  term.term_year(),
                                                  student.lesson_from_name(lesson_name).name,
                                                  lesson_name.upper(),
                                                  term.term_year())
                if sheet["B" + str(k)].value is None:
                    sheet["B" + str(k)].value = student.numara
                    sheet["C" + str(k)].value = student.ad
                else:
                    while sheet["B" + str(k)].value:
                        k += 1
                    sheet["B" + str(k)].value = student.numara
                    sheet["C" + str(k)].value = student.ad

    kurs_wb.save(filename)


# ------------------- #
# For Another Project #
def save_list_to_csv(folder, data_dict):
    lesson_lists = data_dict["lists"]
    students = data_dict["students"]

    for lesson_list in lesson_lists:
        filename = str(lesson_list + ".csv")
        with open(folder + filename, "w+", encoding="utf-8") as file:

            for k in range(len(filename)):
                if filename[k].isdigit():
                    lesson_name = filename[:k]
                    break
            for student in students:
                if lesson_name + student.lesson_from_name(lesson_name).name == filename[:filename.find(".")]:
                    file.write(f"{student.ad},{student.numara}\n")
